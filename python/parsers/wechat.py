import re
from datetime import datetime
from openpyxl import load_workbook

def parse_wechat_bill(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = None
    header_row = 0
    for i, row in enumerate(rows_iter, start=1):
        non_empty = [c for c in row if c not in (None, '')]
        if len(non_empty) >= 8:
            header_row = i
            header = [str(c) if c is not None else '' for c in row]
            break

    if not header:
        return [], '无法定位表头'

    col_map = {}
    for idx, name in enumerate(header):
        if '交易时间' in name:
            col_map['time'] = idx
        elif '交易类型' in name:
            col_map['type'] = idx
        elif '交易对方' in name:
            col_map['counterparty'] = idx
        elif '商品' in name:
            col_map['item'] = idx
        elif '收/支' in name:
            col_map['direction'] = idx
        elif '金额' in name:
            col_map['amount'] = idx
        elif '支付方式' in name:
            col_map['method'] = idx
        elif '当前状态' in name:
            col_map['status'] = idx
        elif '交易单号' in name:
            col_map['external_id'] = idx
        elif '商户单号' in name:
            col_map['merchant_id'] = idx
        elif '备注' in name:
            col_map['note'] = idx

    required = ['time', 'direction', 'amount', 'external_id']
    missing = [r for r in required if r not in col_map]
    if missing:
        return [], f'缺少必要列: {missing}'

    transactions = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
        if not row or not isinstance(row[col_map['time']], datetime):
            continue

        direction_raw = str(row[col_map['direction']]).strip()
        if '支出' in direction_raw:
            direction = 'expense'
        elif '收入' in direction_raw:
            direction = 'income'
        else:
            direction = 'neutral'

        amount = float(row[col_map['amount']]) if row[col_map['amount']] is not None else 0

        external_id = str(row[col_map['external_id']]).strip()
        status = str(row[col_map['status']]).strip() if 'status' in col_map and row[col_map['status']] else ''

        merchant_id = ''
        if 'merchant_id' in col_map and row[col_map['merchant_id']]:
            merchant_id = str(row[col_map['merchant_id']]).strip()

        counterparty = str(row[col_map['counterparty']]).strip() if 'counterparty' in col_map and row[col_map['counterparty']] else ''
        item_desc = str(row[col_map['item']]).strip() if 'item' in col_map and row[col_map['item']] else ''
        payment_method = str(row[col_map['method']]).strip() if 'method' in col_map and row[col_map['method']] else ''
        note = str(row[col_map['note']]).strip() if 'note' in col_map and row[col_map['note']] else ''

        trade_type = str(row[col_map['type']]).strip() if 'type' in col_map else ''

        flow_type = infer_flow_type_wechat(trade_type)

        transactions.append({
            'occurred_at': row[col_map['time']],
            'direction': direction,
            'amount': amount,
            'channel': 'wechat',
            'counterparty': counterparty,
            'item_desc': item_desc,
            'payment_method': payment_method,
            'status': status,
            'external_id': external_id,
            'merchant_id': merchant_id,
            'flow_type': flow_type,
            'note': note,
        })

    wb.close()
    return transactions, None

def infer_flow_type_wechat(trade_type):
    t = trade_type
    if any(k in t for k in ['商户消费', '扫二维码付款', '充值缴费', '红包']):
        return 'consumption'
    elif '转账' in t:
        return 'transfer'
    elif '信用卡还款' in t:
        return 'credit_repay'
    elif any(k in t for k in ['收款', '退款', '转入']):
        return 'income'
    return 'unknown'
